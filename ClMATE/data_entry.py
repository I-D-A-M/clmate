from PyQt4 import QtGui, QtCore
from .helpers import mean, median
from .bases import StdWindow
from .main_interface import OverviewWidget
from .definitions import QBOX_SIZE, grey, white, red, purple, orange, green, dGrey, export_path
from collections import namedtuple
from openpyxl import Workbook
import sqlite3
import time
import sys
import os


class InputWindow(QtGui.QWidget):
    def __init__(self, session_details, CHOSEN_CLASS,
                 CLASS_SIZE, CHOSEN_ASSESSMENTS):
        super(InputWindow, self).__init__()
        self.session_details = session_details
        self.CHOSEN_CLASS = str(CHOSEN_CLASS)
        self.CLASS_SIZE = int(CLASS_SIZE[0])
        self.CHOSEN_ASSESSMENTS = CHOSEN_ASSESSMENTS
        self.grade_dict_list = []
        self.boundary_list = []
        self.col_max_list = []
        self.grade_comp_list = []

        # --  Set up the main window interface
        self.stackLayout = QtGui.QStackedLayout()
        self.scrollWidget = QtGui.QWidget()
        self.scrollWidget.setLayout(self.stackLayout)
        self.scrollArea = QtGui.QScrollArea()
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setWidget(self.scrollWidget)
        self.switcher = QtGui.QComboBox()
        self.mainLayout = QtGui.QVBoxLayout()

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            for a in range(len(self.CHOSEN_ASSESSMENTS)):
                self.CHOSEN_ASSESSMENT = self.CHOSEN_ASSESSMENTS[a]
                a_string = str(self.CHOSEN_ASSESSMENT)
                query = "select count(*) from assessments where aName = ?"
                self.NUM_QUESTIONS = DB.execute(query, (a_string,)).fetchone()[0]
                query = "select qTitle from assessments where aName = ?"
                raw_names = DB.execute(query, (a_string,)).fetchall()
                self.QNAMES = []
                for n in range(len(raw_names)):
                    self.QNAMES.append(str(raw_names[n][0]))
                self.assessment = QtGui.QWidget()
                self.switcher.addItem(self.CHOSEN_ASSESSMENT)
                self.switcher.setCurrentIndex(a)
                self.stackLayout.addWidget(self.assessment)
                grid = QtGui.QGridLayout()
                grid.addWidget(self.createBoundaries(), 0, 0)
                grid.addWidget(self.createButtons(), 3, 0)
                grid.addWidget(self.mainArea(), 0, 1, 4, 8)
                self.assessment.resize(grid.sizeHint())
                self.assessment.setLayout(grid)
        DB.close()

        self.mainLayout.addWidget(self.switcher)
        self.mainLayout.addWidget(self.scrollArea)
        self.setLayout(self.mainLayout)
        self.switcher.connect(self.switcher,
                              QtCore.SIGNAL("currentIndexChanged(int)"),
                              self.tabSwitch)

        # Initialise the first assessment in the stackLayout
        for assessment in range(self.stackLayout.count()):
            self.stackLayout.setCurrentIndex(assessment)
            self.switcher.setCurrentIndex(assessment)
            # This returns the 'main' widget and the elements
            # have to be accessed via the .itemAt method::
            a_layout = self.stackLayout.currentWidget().layout()
            active = a_layout.itemAtPosition(0, 1).widget()
            entry_table = active.layout().itemAt(0).widget()
            entry_table.itemChanged.connect(self.update_summaries)
        # Reset the stack layout to the first assessment
        self.stackLayout.setCurrentIndex(0)
        self.switcher.setCurrentIndex(0)
        self.update_summaries()

    def createBoundaries(self):
        '''
        Create the grade boundary panel for the user to refer to.
        This function also compares the assessment's stated available
        grades against the target grades for the current class and
        disables grade based analysis if there is a mismatch.
        '''
        boundaryBox = QtGui.QWidget()
        bboxlayout = QtGui.QVBoxLayout()

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            query = ("select * from gradeBoundaries where aName = ? "
                     "order by perc desc")
            boundaries = DB.execute(query, (self.CHOSEN_ASSESSMENT,)).fetchall()
        DB.close()

        bHeadings = QtGui.QWidget()
        bHeadingLayout = QtGui.QHBoxLayout()
        bHeadingLayout.addWidget(QtGui.QLabel("Grade"))
        bHeadingLayout.addWidget(QtGui.QLabel("  %  "))
        bHeadings.setLayout(bHeadingLayout)
        bboxlayout.addWidget(bHeadings)
        for b in boundaries:
            row = QtGui.QWidget()
            rowlayout = QtGui.QHBoxLayout()
            grade = QtGui.QLabel(str(b[2]))
            perc = QtGui.QLabel(str(b[3]))
            rowlayout.addWidget(grade)
            rowlayout.addWidget(perc)
            row.setLayout(rowlayout)
            rowlayout.setSpacing(0)
            rowlayout.setContentsMargins(0, 0, 0, 0)
            bboxlayout.addWidget(row)
        bboxlayout.setSpacing(0)
        bboxlayout.setContentsMargins(0, 0, 0, 0)
        boundaryBox.setLayout(bboxlayout)

        # Store assessment grades for reference.
        grade_order = []
        for b in range(len(boundaries)):
            grade_order.append((str(boundaries[b][2]), len(boundaries) - b))
        # ('N', 0) allows for 'N/A' to be set (value = 0) for pupils
        # with no available target
        # NOTE:: This is 'N' not 'N/A' as targets are filtered for
        #        split grades if they are more than 2 characters
        #        long (this allows for A* / B- etc).
        grade_order.append(('N', 0))
        grade_dict = dict(grade_order)
        self.boundary_list.append(boundaries)
        self.grade_dict_list.append(grade_dict)

        return boundaryBox

    def createButtons(self):
        buttonBox = QtGui.QGroupBox()
        paperButton = QtGui.QPushButton("&Paper")
        paperButton.clicked.connect(self.openpaper)
        MSButton = QtGui.QPushButton("&Mark Scheme")
        MSButton.clicked.connect(self.openMS)
        overviewButton = QtGui.QPushButton("&Overview")
        overviewButton.clicked.connect(self.quick_overview)
        saveButton = QtGui.QPushButton("&Save")
        saveButton.clicked.connect(self.commit_to_DB)
        exitButton = QtGui.QPushButton("Save and &Exit")
        exitButton.clicked.connect(self.replace_main_widget)

        vbox = QtGui.QVBoxLayout()
        vbox.addWidget(paperButton)
        vbox.addWidget(MSButton)
        vbox.addWidget(overviewButton)
        vbox.addWidget(saveButton)
        vbox.addWidget(exitButton)
        buttonBox.setLayout(vbox)
        return buttonBox

    def mainArea(self):
        # This table is NOT a live connection to the database::
        # data is read in to populate the view and then stripped out on save.
        main = QtGui.QWidget()
        vbox = QtGui.QVBoxLayout()
        grade_dict = self.grade_dict_list[self.switcher.currentIndex()]
        GRADE_COMP_ENABLED = True

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            # Locate pupil names and store as a list of strings.
            query = "select name from cohort where teaching_set = ?"
            NAMES = DB.execute(query, (self.CHOSEN_CLASS,)).fetchall()
            self.name_list = []
            for p in range(len(NAMES)):
                n = str(NAMES[p][0])
                self.name_list.append(n)
            # Hard coded at present but ideally this should be a function
            # that can be set by the user when they set up a class / course.
            query = "select yrGroup from staffing where teaching_set = ?"
            yrGroup = DB.execute(query, (self.CHOSEN_CLASS,)).fetchone()[0]
            if yrGroup == 'GCSE':
                query = "select FFT from cohort where teaching_set = ?"
                TARGETS = DB.execute(query, (self.CHOSEN_CLASS,)).fetchall()
                targetName = 'FFT'
            elif yrGroup == 'AS':
                query = "select ASAlps from cohort where teaching_set = ?"
                TARGETS = DB.execute(query, (self.CHOSEN_CLASS,)).fetchall()
                targetName = 'ALPs'
            elif yrGroup == 'A2':
                query = "select A2Alps from cohort where teaching_set = ?"
                TARGETS = DB.execute(query, (self.CHOSEN_CLASS,)).fetchall()
                targetName = 'ALPs'
            else:
                QtGui.QMessageBox.question(
                    self,
                    "Warning!",
                    "Unable to find year group for current class.")
                return -1
        DB.close()

        # Compare pupil targets to the assessment grade dictionary.
        # This is where GRADE_COMP_ENABLED will be set to False if
        # a key error is raised in trying to locate a target.
        target_list = []
        target_score_list = []
        for t in range(len(TARGETS)):
            g = str(TARGETS[t][0])
            # Input format from openpyxl defaults to 'None' if a
            # cell is empty so we need to overwrite it now.
            if g == 'None':
                g = 'N/A'
            target_list.append(g)
            # Some targets will be split over two possible values:
            # this would cause a problem as the grade boundary dictionary
            # keys would raise an error. ClMATE treats these with a
            # positional 'value' of (upper grade - 0.5) under the assumption
            # that a split target will always be between adjacent grades/levels.
            split_target = False
            if len(g) > 2:
                g = g[0]
                split_target = True
            try:
                val = grade_dict[g]
                if split_target:
                    val -= 0.5
                target_score_list.append(val)
            except KeyError:
                GRADE_COMP_ENABLED = False
                val = 0
                target_score_list.append(val)
        self.grade_comp_list.append(GRADE_COMP_ENABLED)
        # Only run grade comparison code if grades are valid for
        # the stated boundaries with the current assessment.
        median_score = int(round(median(target_score_list)))
        mean_score = int(round(mean(target_score_list)))
        dict_items_for_mean = grade_dict.items()
        dict_items_for_median = grade_dict.items()
        # -- Try to locate the mean and median target grades for the class.
        try:
            for grade, val in dict_items_for_median:
                if val == median_score:
                    median_target = grade
        except KeyError:
            median_target = 'N/A'
        try:
            for grade, val in dict_items_for_mean:
                if val == mean_score:
                    mean_target = grade
        except KeyError:
            mean_target = 'N/A'
        stats_list = [mean_target, median_target, '']
        # The main area is made up of a data entry table with pupil level
        # totals etc and a summary table with class level question breakdown.
        questions = QtGui.QTableWidget(self.CLASS_SIZE, self.NUM_QUESTIONS + 6)
        summary = QtGui.QTableWidget(3, self.NUM_QUESTIONS + 6)
        summary.setFixedHeight(130)
        # -- Get rid of the row numbering to clean up the UI
        blank_labels = ['' for x in range(self.CLASS_SIZE)]
        for q in range(self.NUM_QUESTIONS + 6):
            questions.setColumnWidth(q, QBOX_SIZE)
            summary.setColumnWidth(q, QBOX_SIZE)
        questions.setVerticalHeaderLabels(blank_labels)
        # -- Define which cells are editable, initial content and colourschemes
        for r in range(self.CLASS_SIZE):
            namecell = QtGui.QTableWidgetItem()
            namecell.setBackgroundColor(grey)
            namecell.setFlags(QtCore.Qt.NoItemFlags)
            namecell.setText(self.name_list[r])
            namecell.setTextColor(white)
            questions.setItem(r, 0, namecell)

            targetcell = QtGui.QTableWidgetItem()
            targetcell.setBackgroundColor(grey)
            targetcell.setFlags(QtCore.Qt.NoItemFlags)
            targetcell.setText(target_list[r])
            targetcell.setTextColor(white)
            questions.setItem(r, 1, targetcell)

            totalcell = QtGui.QTableWidgetItem()
            totalcell.setBackgroundColor(grey)
            totalcell.setFlags(QtCore.Qt.NoItemFlags)
            totalcell.setTextColor(white)
            questions.setItem(r, self.NUM_QUESTIONS + 2, totalcell)

            percentcell = QtGui.QTableWidgetItem()
            percentcell.setBackgroundColor(grey)
            percentcell.setFlags(QtCore.Qt.NoItemFlags)
            percentcell.setTextColor(white)
            questions.setItem(r, self.NUM_QUESTIONS + 3, percentcell)

            gradecell = QtGui.QTableWidgetItem()
            gradecell.setBackgroundColor(grey)
            gradecell.setFlags(QtCore.Qt.NoItemFlags)
            gradecell.setTextColor(white)
            questions.setItem(r, self.NUM_QUESTIONS + 4, gradecell)

            relativecell = QtGui.QTableWidgetItem()
            relativecell.setBackgroundColor(grey)
            relativecell.setFlags(QtCore.Qt.NoItemFlags)
            relativecell.setTextColor(white)
            if not GRADE_COMP_ENABLED:
                relativecell.setText('##')
            questions.setItem(r, self.NUM_QUESTIONS + 5, relativecell)

        # -- Initialise the summary panel
        sum_titles = ['Mean', 'Median', '% marks obtained']
        for s in range(3):
            statcell = QtGui.QTableWidgetItem()
            statcell.setBackgroundColor(grey)
            statcell.setFlags(QtCore.Qt.NoItemFlags)
            statcell.setText(sum_titles[s])
            statcell.setTextColor(white)
            summary.setItem(s, 0, statcell)

            meanMed = QtGui.QTableWidgetItem()
            meanMed.setBackgroundColor(grey)
            meanMed.setFlags(QtCore.Qt.NoItemFlags)
            meanMed.setText(stats_list[s])
            meanMed.setTextColor(white)
            summary.setItem(s, 1, meanMed)

            totalcell = QtGui.QTableWidgetItem()
            totalcell.setBackgroundColor(grey)
            totalcell.setFlags(QtCore.Qt.NoItemFlags)
            totalcell.setTextColor(white)
            summary.setItem(s, self.NUM_QUESTIONS + 2, totalcell)

            percentcell = QtGui.QTableWidgetItem()
            percentcell.setBackgroundColor(grey)
            percentcell.setFlags(QtCore.Qt.NoItemFlags)
            percentcell.setTextColor(white)
            summary.setItem(s, self.NUM_QUESTIONS + 3, percentcell)

            gradecell = QtGui.QTableWidgetItem()
            gradecell.setBackgroundColor(grey)
            gradecell.setFlags(QtCore.Qt.NoItemFlags)
            gradecell.setTextColor(white)
            summary.setItem(s, self.NUM_QUESTIONS + 4, gradecell)

            relativecell = QtGui.QTableWidgetItem()
            relativecell.setBackgroundColor(grey)
            relativecell.setFlags(QtCore.Qt.NoItemFlags)
            relativecell.setTextColor(white)
            if not GRADE_COMP_ENABLED:
                relativecell.setText('##')
            summary.setItem(s, self.NUM_QUESTIONS + 5, relativecell)

            for q in range(self.NUM_QUESTIONS):
                qcell = QtGui.QTableWidgetItem()
                qcell.setBackgroundColor(white)
                qcell.setTextColor(dGrey)
                qcell.setFlags(QtCore.Qt.NoItemFlags)
                summary.setItem(s, q + 2, qcell)
        # Resise the name and target columns for both panels
        questions.resizeColumnToContents(0)
        summary.setColumnWidth(0, questions.columnWidth(0))
        questions.setColumnWidth(1, 60)
        summary.setColumnWidth(1, 60)
        # Set the column headers for both panels
        qHeader_list = ['Name', targetName]
        sHeader_list = ['', '']
        for n in range(self.NUM_QUESTIONS):
            qHeader_list.append(str(n + 1))
            sHeader_list.append(str(n + 1))
        qHeader_list.append('Total')
        sHeader_list.append('Total')
        qHeader_list.append('  %  ')
        sHeader_list.append('  %  ')
        qHeader_list.append('Grade')
        sHeader_list.append('Grade')
        qHeader_list.append(u'\u21c5')
        sHeader_list.append(u'\u21c5')

        # Resize the last three columns so that they are always in line
        for n in range(2, 6):
            questions.setColumnWidth(self.NUM_QUESTIONS + n, 55)
            summary.setColumnWidth(self.NUM_QUESTIONS + n, 55)

        # Find the total marks available for each question and
        # store them for use in update_summaries()
        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            col_maximums = []
            for q in range(self.NUM_QUESTIONS):
                query = "select qMark from assessments where aName = ? and qNum = ?"
                qMark = DB.execute(query, (self.CHOSEN_ASSESSMENT, q + 1)).fetchone()
                col_maximums.append(qMark[0])
        DB.close()

        total_marks = sum(col_maximums)
        col_maximums.append(total_marks)
        col_maximums.append(100)
        self.col_max_list.append(col_maximums)

        # -- Set column and row headings [including tooltip question names]
        questions.setVerticalHeaderLabels(blank_labels)
        questions.setHorizontalHeaderLabels(qHeader_list)
        summary.setVerticalHeaderLabels(['', '', ''])
        summary.setHorizontalHeaderLabels(sHeader_list)
        for q in range(self.NUM_QUESTIONS):
            questions.horizontalHeaderItem(q + 2).setToolTip(
                self.QNAMES[q] +
                "  ( /" + str(col_maximums[q]) + ")")
            summary.horizontalHeaderItem(q + 2).setToolTip(
                self.QNAMES[q] +
                "  ( /" + str(col_maximums[q]) + ")")

        # Populate the questions panel with prior data if found
        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            aName = self.CHOSEN_ASSESSMENT
            teaching_set = self.CHOSEN_CLASS
            query = "select pMark from results where aName = ? and teaching_set = ?"
            already_entry = DB.execute(query, (aName, teaching_set)).fetchone()
            # -- Default to all cell values == 0
            if already_entry is None:
                for row in range(self.CLASS_SIZE):
                    for col in range(self.NUM_QUESTIONS):
                        questions.setItem(row,
                                          col + 2,
                                          QtGui.QTableWidgetItem("%d" % 0))
            else:
                # -- Read in stored data from the last data entry session.
                query = ("select pMark from results where "
                         "teaching_set = ? and aName = ?")
                class_results = DB.execute(query,
                                           (self.CHOSEN_CLASS, aName)).fetchall()
                for row in range(self.CLASS_SIZE):
                    # This should slice the results list into pupil
                    # chunks with each one ordered by qNum
                    lower = (row * self.NUM_QUESTIONS)
                    upper = ((row + 1) * self.NUM_QUESTIONS)
                    current_pupil_results = class_results[lower: upper]
                    for col in range(self.NUM_QUESTIONS):
                        pMark = int(current_pupil_results[col][0])
                        questions.setItem(row,
                                          col + 2,
                                          QtGui.QTableWidgetItem("%d" % pMark))
        DB.close()

        # -- Clipboard management details
        self.clipboard = QtGui.QApplication.instance().clipboard()
        # -- Set the keyboard shortcuts ctrl+c for copy and ctrl+v fo r paste
        paste = QtGui.QShortcut(QtGui.QKeySequence('Ctrl+v'), questions)
        paste.activated.connect(self.handlePaste)
        copy = QtGui.QShortcut(QtGui.QKeySequence('Ctrl+c'), questions)
        copy.activated.connect(self.handleCopy)
        save_to_DB = QtGui.QShortcut(QtGui.QKeySequence('Ctrl+s'), questions)
        save_to_DB.activated.connect(self.commit_to_DB)
        # -- Populate the main widget and return
        vbox.addWidget(questions)
        vbox.addWidget(summary)
        vbox.setSpacing(0)
        vbox.setContentsMargins(0, 0, 0, 0)
        main.setLayout(vbox)
        return main

    def tabSwitch(self):
        '''
        This function will take the current tab index and match
        the central widget index to it.
        '''
        self.stackLayout.setCurrentIndex(self.switcher.currentIndex())
        self.update_summaries()

    def handleCopy(self):
        copied_text = ''
        layout = self.stackLayout.currentWidget().layout()
        active = layout.itemAtPosition(0, 1).widget()
        questions = active.layout().itemAt(0).widget()
        sRange = questions.selectedRanges()[0]
        for row in range(sRange.rowCount()):
            if row > 0:
                copied_text += '\n'
            for col in range(sRange.columnCount()):
                if col > 0:
                    copied_text += '\t'
                selection = questions.item(sRange.topRow() + row,
                                           sRange.leftColumn() + col)
                copied_text += str(selection.text())
        self.clipboard.setText(copied_text)

    def handlePaste(self):
        # NOTE:: this needs to do bounds checking to ensure
        # that we don't try to paste over and statistical cells!
        # Will need to selectively drop the rest of the current
        # row / selection if out of bounds.
        layout = self.stackLayout.currentWidget().layout()
        active = layout.itemAtPosition(0, 1).widget()
        questions = active.layout().itemAt(0).widget()
        sRange = questions.selectedRanges()[0]
        clipboard_text = self.clipboard.text()
        rows = clipboard_text.split('\n')
        numRows = len(rows)
        numColumns = len(rows[0].split('\t'))
        ROW, COL = sRange.topRow(), sRange.leftColumn()

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            CHOSEN_ASSESSMENT = str(self.switcher.currentText())
            query = "select count(*) from assessments where aName = ?"
            NUM_QUESTIONS = DB.execute(query, (CHOSEN_ASSESSMENT, )).fetchone()[0]
        DB.close()

        # -- Only paste if the selection will contain ALL of the data
        # in the clipboard. NOTE: There is a 2 column offset on the
        # Q numbers due to the name and target fields.
        # -- Remove the last copied row if it is a null string
        # [spreadsheets append a null-string to mark the end
        # of the copy selection]
        if rows[-1] == '':
            rows = rows[:-1]
            numRows = len(rows)
        if ((ROW + numRows <= self.CLASS_SIZE) and
           (COL + numColumns <= NUM_QUESTIONS + 2)):
            active.blockSignals(True)
            for r in enumerate(rows):
                columns = r[1].split('\t')
                for c in enumerate(columns):
                    row = sRange.topRow() + r[0]
                    column = sRange.leftColumn() + c[0]
                    questions.item(row, column).setText(c[1])
            self.update_summaries()
            active.blockSignals(False)
        # -- Warn the user that the paste action has been aborted
        else:
            QtGui.QMessageBox.question(
                self,
                'Warning!',
                "Copied values will not fit in the selected range.")
            return -1

    def replace_main_widget(self):
        error_catch = self.commit_to_DB()
        if error_catch == 'no errors':
            overviewWidget = OverviewWidget(self.session_details)
            self.session_details["main_window"].setCentralWidget(overviewWidget)

    def openpaper(self):
        # locate assessment paper
        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            CHOSEN_ASSESSMENT = str(self.switcher.currentText())
            query = "select paper from assessmentDocs where aName = ?"
            paper_query = DB.execute(query, (CHOSEN_ASSESSMENT,)).fetchone()[0]
        DB.close()

        self.paper_path = str(paper_query)
        os.startfile(self.paper_path)

    def openMS(self):
        # locate MS path
        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            CHOSEN_ASSESSMENT = str(self.switcher.currentText())
            query = "select MS from assessmentDocs where aName = ?"
            ms_query = DB.execute(query, (CHOSEN_ASSESSMENT,)).fetchone()[0]
            self.MS_path = str(ms_query)
        DB.close()
        os.startfile(self.MS_path)

    def quick_overview(self):
        '''
        This should specify:    % distribution of grades (splits)
                                SEN/PP vs class, cohort and other SEN/PP
                                % on / above / below target
                                pupil ranking within class
                                [add to this!]
        '''
        # Retrieve the current active questions widget from the grid layout
        layout = self.stackLayout.currentWidget().layout()
        active = layout.itemAtPosition(0, 1).widget()
        questions = active.layout().itemAt(0).widget()
        # Identify assessment properties based on which assessment is
        # currently active in the InpuWindow widget.
        boundaries = self.boundary_list[self.switcher.currentIndex()]
        col_maximums = self.col_max_list[self.switcher.currentIndex()]
        # col_maximums = [(Q totals...), total, 100]
        GRADE_COMP_ENABLED = self.grade_comp_list[self.switcher.currentIndex()]
        CHOSEN_ASSESSMENT = str(self.switcher.currentText())

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            query = "select count(*) from assessments where aName = ?"
            NUM_QUESTIONS = DB.execute(query, (CHOSEN_ASSESSMENT, )).fetchone()[0]
        DB.close()

        ####################################################
        # -- Find the top/bottom three questions by %marks #
        ####################################################
        qStatsList = []
        stat = namedtuple("QuesitonsStats", "Num Mean Median Perc")
        for Q in range(NUM_QUESTIONS):
            class_marks = []
            for p in range(self.CLASS_SIZE):
                qNum = Q + 1
                class_marks.append(int(questions.item(p, Q + 2).text()))
            qMean = mean(class_marks)
            qMedian = median(class_marks)
            total_q_mark = col_maximums[Q] * self.CLASS_SIZE
            qPerc = int(round(float(sum(class_marks)) / (total_q_mark) * 100))
            q_details = stat(qNum, qMean, qMedian, qPerc)
            qStatsList.append(q_details)
        # -- create a sorted version of the list by percentage of marks achieved
        sorted_by_perc = sorted(qStatsList, key=lambda Q: Q.Perc)
        top3 = sorted_by_perc[-3:]
        top3.reverse()
        bottom3 = sorted_by_perc[:3]
        bottom3.reverse()

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            top3Qdetails = ''
            for q in top3:
                query = "select qTitle from assessments where aName = ?"
                qname = DB.execute(query,
                                   (CHOSEN_ASSESSMENT,)).fetchall()[q.Num - 1][0]
                qSummary = '\t' + str(q.Num) + ') ' + qname + '  ' + str(q.Perc) + '%\n'
                top3Qdetails += qSummary
            bottom3Qdetails = ''
            for q in bottom3:
                query = "select qTitle from assessments where aName = ?"
                qname = DB.execute(query,
                                   (CHOSEN_ASSESSMENT,)).fetchall()[q.Num - 1][0]
                qSummary = '\t' + str(q.Num) + ') ' + qname + '  ' + str(q.Perc) + '%\n'
                bottom3Qdetails += qSummary
        DB.close()

        # -- Set summary message
        topBottomQs = ("Top three questions by % marks achieved:\n" +
                       top3Qdetails +
                       "\nBottom three questions by % marks achieved:\n" +
                       bottom3Qdetails)

        #########################################
        # -- Find %A*, A*-A, A*-C or equivalent #
        #########################################
        if GRADE_COMP_ENABLED:
            # -- Count the occurances of each grade and store them in a tuple
            pGrades = []
            gradeCount = []
            for p in range(self.CLASS_SIZE):
                grade = questions.item(p, NUM_QUESTIONS + 4).text()
                pGrades.append(grade)
            for g in boundaries:
                gradeCount.append(pGrades.count(g[2]))
            gradeSplit = '% Grade splits for this assessment:\n'
            perc = float("{0:.1f}".format(gradeCount[0] / self.CLASS_SIZE * 100))
            perc = str(perc)
            gradeSplit = (gradeSplit +
                          '\t' + boundaries[0][2] +
                          ': ' +
                          perc + '%\n')
            sumSoFar = gradeCount[0]
            for g in enumerate(boundaries):
                if g[1][2] != boundaries[0][2]:
                    sumSoFar += gradeCount[g[0]]
                    perc = float("{0:.1f}".format(sumSoFar / self.CLASS_SIZE * 100))
                    perc = str(perc)
                    gradeSplit = (gradeSplit +
                                  '\t' + boundaries[0][2] +
                                  '-' + g[1][2] +
                                  ': ' +
                                  perc + '%\n')
        gradeSplit = '% Grade splits have been disabled for this assessment.\n'
        # -- Build and populate the overview
        global overview
        overview = StdWindow()
        overviewLayout = QtGui.QGridLayout()
        overviewLayout.addWidget(QtGui.QLabel(topBottomQs), 0, 0)
        overviewLayout.addWidget(QtGui.QLabel(gradeSplit), 1, 0)
        save_button = QtGui.QPushButton('save to xlsx')
        save_button.clicked.connect(self.export_to_xlsx)
        overviewLayout.addWidget(save_button, 2, 0)
        overview.setLayout(overviewLayout)
        overview.resize(500, 300)
        overview.initUI()
        overview.setWindowTitle('Overview for ' +
                                str(self.switcher.currentText()))

    def commit_to_DB(self):
        username = self.session_details["username"]
        with open('clmate_log.txt', 'a') as log:
            print("    {} starting save of data for {} at: {}".format(
                  username,
                  self.CHOSEN_CLASS,
                  time.ctime()),
                  file=log)

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            layout = self.stackLayout.currentWidget().layout()
            active = layout.itemAtPosition(0, 1).widget()
            questions = active.layout().itemAt(0).widget()
            aName = str(self.switcher.currentText())
            query = "select aID from assignedTests where aName = ?"
            aID = DB.execute(query, (aName,)).fetchone()[0]
            aID = int(aID)
            query = "select count(*) from assessments where aName = ?"
            NUM_QUESTIONS = DB.execute(query, (aName, )).fetchone()[0]
            teaching_set = self.CHOSEN_CLASS

            results = list()
            query = "select UPN from cohort where name = ? and teaching_set = ?"
            UPN = DB.execute(query, (self.name_list[0], teaching_set)).fetchone()[0]
            query = ("select pMark from results where UPN = ? "
                     "and aName = ? and qNum = ?")
            already_entries = DB.execute(query, (UPN, aName, 1)).fetchone()

            for p in range(self.CLASS_SIZE):
                start = time.time()
                query = "select UPN from cohort where name = ? and teaching_set = ?"
                UPN = DB.execute(query,
                                 (self.name_list[p], teaching_set)).fetchone()[0]
                for q in range(NUM_QUESTIONS):
                    qNum = q + 1
                    pMark = int(questions.item(p, q + 2).text())
                    # Test and block saving if errors were made in data entry
                    if pMark == 999999:
                        QtGui.QMessageBox.question(
                            self,
                            'Warning!',
                            "Errors were detected in the data you have entered."
                            "Please re-check any red highlighted cells.")
                        results = None
                        return 'errors found in data'
                    else:
                        results.append((pMark, aID, UPN, teaching_set, aName, qNum))

            DB.execute("PRAGMA synchronous = OFF")
            DB.execute("PRAGMA journal_mode = OFF")
            if already_entries:
                query = ("update results set pMark = ? where aID = ? and UPN = ? "
                         "and teaching_set = ? and aName = ? and qNum = ?")
                DB.executemany(query, results)
            else:
                query = ("insert into results (pMark, aID, UPN, teaching_set, "
                         "aName, qNum) values (?,?,?,?,?,?)")
                DB.executemany(query, results)
        DB.close()

        fin = time.time() - start
        with open('clmate_log.txt', 'a') as log:
            print("    {} completed save operation in {} seconds.".format(
                username, fin), file=log)
        return 'no errors'

    def export_to_xlsx(self, breakdown_included=False):
        '''
        This should take a [questions] widget and rip out the details
        for saving to excel. It should also implement the {Breakdown}
        analysis and save that as a second sheet.
        '''
        # -- Locate required fields
        layout = self.stackLayout.currentWidget().layout()
        active = layout.itemAtPosition(0, 1).widget()
        questions = active.layout().itemAt(0).widget()
        aName = str(self.switcher.currentText())

        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            query = "select count(*) from assessments where aName = ?"
            NUM_QUESTIONS = DB.execute(query, (str(aName),)).fetchone()[0]
        DB.close()

        teaching_set = self.CHOSEN_CLASS
        validCName = ''
        for ch in teaching_set:
            if ch == '/':
                ch = '_'
                validCName += ch
            elif ch == ' ':
                ch = '_'
                validCName += ch
            else:
                validCName += ch
        # This needs to be set based on the OS
        OS = sys.platform
        if OS == 'linux' or OS == 'linux2':
            title = validCName + '_' + aName + '_' + time.strftime("%d-%m-%Y") + '.xlsx'
        elif OS == 'win32':
            # Default path for export will need to be set
            # in the first time wizard
            title = (export_path +
                     validCName +
                     '_' + aName +
                     '_' + time.strftime("%d-%m-%Y") +
                     '.xlsx')
        else:
            QtGui.QMessageBox.question(self, 'Error!', "Could not identify OS.")
            return -1
        # Block saving if there are errors in the data
        for p in range(self.CLASS_SIZE):
            for q in range(NUM_QUESTIONS):
                pMark = int(questions.item(p, q + 2).text())
                # Test and block saving if errors were made in data entry
                if pMark == 999999:
                    QtGui.QMessageBox.question(
                        self,
                        'Warning!',
                        "Errors were detected in the data you have entered."
                        "Please re-check any red highlighted cells.")
                    return -1
        # This uses openpyxl
        wb = Workbook()
        # This will create a new sheet --> ws = wb.create_sheet()
        ws = wb.active
        ws.title = validCName
        # -- Set up column headings
        headings = ['UPN', 'Pupil name', 'Target']
        for n in range(NUM_QUESTIONS):
            headings.append(str(n + 1))
        headings.append('Total')
        headings.append('  %  ')
        headings.append('Grade')
        headings.append('Relative')
        for col in range(NUM_QUESTIONS + 7):
            heading = ws.cell(row=1, column=(col + 1))
            heading.value = headings[col]

        # -- Rip data
        DBname = self.session_details["DBname"]
        with sqlite3.connect(DBname) as DB:
            for r in range(self.CLASS_SIZE):
                upn_val = DB.execute("select UPN from cohort where name = ?",
                                     (self.name_list[r],)).fetchone()[0]
                upn = ws.cell(row=(r + 2), column=1)
                upn.value = upn_val
                for c in range(NUM_QUESTIONS + 5):
                    cell_contents = str(questions.item(r, c).text())
                    excel_cell = ws.cell(row=(r + 2), column=(c + 2))
                    excel_cell.value = cell_contents
                rel = str(questions.item(r, NUM_QUESTIONS + 5).text()[1:])
                excel_cell = ws.cell(row=(r + 2), column=(NUM_QUESTIONS + 7))
                excel_cell.value = rel
        DB.close()

        wb.save(title)
        QtGui.QMessageBox.question(
            self,
            'Success',
            "Please check the ClMATE output directory "
            "for your spreadsheet.")
        global overview
        overview.close()

    def update_summaries(self, *args):
        '''
        Handle user input error detection, conditional formatting,
        updating of summary statistics and calculation of pupil grades.
        NOTE:: There is variable execution based on assessment GRADE_COMP_ENABLED
        '''
        # Retrieve the current active questions widget from the grid layout.
        active = self.stackLayout.currentWidget().layout().itemAtPosition(0, 1).widget()
        questions = active.layout().itemAt(0).widget()
        summary = active.layout().itemAt(1).widget()

        # Prevent further signals being sent while updating the table
        # in order to stop our changes triggering another call.
        questions.blockSignals(True)

        # Identify assessment properties based on which assessment is
        # currently active in the InpuWindow widget.
        grade_dict = self.grade_dict_list[self.switcher.currentIndex()]
        boundaries = self.boundary_list[self.switcher.currentIndex()]
        col_maximums = self.col_max_list[self.switcher.currentIndex()]
        # NOTE: col_maximums = [(Q totals...), total, 100]
        # NOTE: This may need to be changed to -1 if I remove the 100 from the end of colmax
        NUM_QUESTIONS = len(col_maximums) - 2

        if len(args) == 0:
            for ROW in range(self.CLASS_SIZE):
                for q in range(NUM_QUESTIONS):
                    COL = q + 2
                    self.validate_and_colour(ROW, COL, questions, col_maximums, NUM_QUESTIONS)
                self.compute_pupil_stats(ROW, questions, col_maximums, NUM_QUESTIONS, boundaries, grade_dict)
        else:
            # Only accept triggers from user editable cells.
            if args[0].column() < NUM_QUESTIONS + 2:
                ROW, COL = args[0].row(), args[0].column()
                self.validate_and_colour(ROW, COL, questions, col_maximums, NUM_QUESTIONS)
                self.compute_pupil_stats(ROW, questions, col_maximums, NUM_QUESTIONS, boundaries, grade_dict)
            else:
                return
        self.compute_summary_stats(ROW, questions, summary, col_maximums, NUM_QUESTIONS, boundaries, grade_dict)

        # Re-enable cell signalling for future updates now that we have finished.
        questions.blockSignals(False)

    def validate_and_colour(self, ROW, COL, questions, col_maximums, NUM_QUESTIONS):
        # Test for invalid entry and then colour.
        if str(questions.item(ROW, COL).text()).isdigit():
            pMark = int(questions.item(ROW, COL).text())
        else:
            questions.setItem(ROW, COL, QtGui.QTableWidgetItem("%s" % '999999'))
            questions.item(ROW, COL).setBackgroundColor(red)
            return

        if (pMark > col_maximums[COL - 2]) or (pMark < 0):
            questions.setItem(ROW, COL, QtGui.QTableWidgetItem("%s" % '999999'))
            questions.item(ROW, COL).setBackgroundColor(red)
            return

        if pMark >= (col_maximums[COL - 2]) * 0.9:
            questions.item(ROW, COL).setBackgroundColor(green)
        elif pMark >= (col_maximums[COL - 2]) * 0.6:
            questions.item(ROW, COL).setBackgroundColor(orange)
        else:
            questions.item(ROW, COL).setBackgroundColor(purple)

    def compute_pupil_stats(self, ROW, questions, col_maximums, NUM_QUESTIONS, boundaries, grade_dict):
        pupil_marks = []
        total_marks = col_maximums[-2]
        for Q in range(NUM_QUESTIONS):
            pMark = int(questions.item(ROW, Q + 2).text())
            if pMark != 999999:
                pupil_marks.append(pMark)
        pupil_total = sum(pupil_marks)
        questions.item(ROW, NUM_QUESTIONS + 2).setText(str(pupil_total))
        pupil_perc = int(round(float(pupil_total) / total_marks * 100))
        questions.item(ROW, NUM_QUESTIONS + 3).setText(str(pupil_perc))

        on_target_sum = []
        # Look through the grade boundaries for the assessment
        # and use the pupil's current percentage to assign a grade.
        for b in boundaries:
            grade = str(b[2])
            perc = b[3]
            if pupil_perc >= perc:
                questions.item(ROW, NUM_QUESTIONS + 4).setText(grade)
                # Some targets will be split over two values:
                # this would cause a problem as the grade boundary
                # dictionary keys would raise an error.
                grade_score = grade_dict[grade]
                g = questions.item(ROW, 1)
                single_target_grade = str(g.text())
                split_target = False
                try:
                    target_score = grade_dict[single_target_grade]
                except KeyError:
                    try:
                        split_target = True
                        single_target_grade = single_target_grade[0]
                        target_score = grade_dict[single_target_grade]
                    except KeyError:
                        target_score = 0
                # NOTE: Treating these with a positional 'value' of (upper grade - 0.5) under
                # the assumption that a split target will be between adjacent grades/levels.
                if split_target:
                    target_score -= 0.5

                if single_target_grade == 'N':
                    on_target = 0
                else:
                    # 0 for on target, +ve for above and -ve for below.
                    on_target = grade_score - target_score
                on_target_sum.append(on_target)

                # The following unicode values denote arrows for indicating pupil attainment
                # relative to target. NOTE: See http://www.alanwood.net/unicode/index.html
                if on_target == 0 or single_target_grade == 'N':
                    questions.item(ROW, NUM_QUESTIONS + 5).setText(
                        u'\u2192' + '  ' + str(on_target))
                elif on_target < 0:
                    questions.item(ROW, NUM_QUESTIONS + 5).setText(
                        u'\u2198' + '  ' + str(on_target))
                else:
                    questions.item(ROW, NUM_QUESTIONS + 5).setText(
                        u'\u2197' + '  ' + str(on_target))
                break

    def compute_summary_stats(self, ROW, questions, summary, col_maximums, NUM_QUESTIONS, boundaries, grade_dict):
        # --Update summary panel regardless of caller.
        class_grade_sum = []
        on_target_sum = []
        for p in range(self.CLASS_SIZE):
            # -- Find class grade total
            # -- Calculate and assign pupil total and percentage values.
            pupil_grade = questions.item(ROW, NUM_QUESTIONS + 4).text()
            grade_score = grade_dict[pupil_grade]
            class_grade_sum.append(grade_score)
            g = questions.item(ROW, 1)
            single_target_grade = str(g.text())
            split_target = False
            try:
                target_score = grade_dict[single_target_grade]
            except KeyError:
                try:
                    split_target = True
                    single_target_grade = single_target_grade[0]
                    target_score = grade_dict[single_target_grade]
                except KeyError:
                    target_score = 0
            # NOTE: Treating these with a positional 'value' of (upper grade - 0.5) under
            # the assumption that a split target will be between adjacent grades/levels.
            if split_target:
                target_score -= 0.5

            if single_target_grade == 'N':
                on_target = 0
            else:
                # 0 for on target, +ve for above and -ve for below.
                on_target = grade_score - target_score
            on_target_sum.append(on_target)

            # -- Calculate class level statistics
            if median(on_target_sum) < 0:
                summary.item(1, NUM_QUESTIONS + 5).setText(
                    u'\u2198' + '  ' + str(median(on_target_sum)))
            elif median(on_target_sum) == 0:
                summary.item(1, NUM_QUESTIONS + 5).setText(
                    u'\u2192' + '  ' + str(median(on_target_sum)))
            else:
                summary.item(1, NUM_QUESTIONS + 5).setText(
                    u'\u2197' + '  ' + str(median(on_target_sum)))
            if mean(on_target_sum) < 0:
                summary.item(0, NUM_QUESTIONS + 5).setText(
                    u'\u2198' + '  ' + str(int(mean(on_target_sum))))
            elif mean(on_target_sum) == 0:
                summary.item(0, NUM_QUESTIONS + 5).setText(
                    u'\u2192' + '  ' + str(int(mean(on_target_sum))))
            else:
                summary.item(0, NUM_QUESTIONS + 5).setText(
                    u'\u2197' + '  ' + str(int(mean(on_target_sum))))

        # -- Caclulate question level statistics
        for Q in range(NUM_QUESTIONS + 2):
            class_marks = []
            for p in range(self.CLASS_SIZE):
                class_marks.append(int(questions.item(p, Q + 2).text()))
            qMean = mean(class_marks)
            qMedian = median(class_marks)
            total_q_mark = (col_maximums[Q] * self.CLASS_SIZE)
            qPerc = int(round(float(sum(class_marks)) / total_q_mark * 100))
            summary.item(0, Q + 2).setText(str(qMean))
            summary.item(1, Q + 2).setText(str(qMedian))
            summary.item(2, Q + 2).setText(str(qPerc))

        # Convert mean and median grade values to ints
        # and look them up in the grade dictionary.
        class_perc = float(summary.item(0, NUM_QUESTIONS + 3).text())
        for b in boundaries:
            grade = str(b[2])
            perc = b[3]
            if class_perc >= perc:
                summary.item(0, NUM_QUESTIONS + 4).setText(grade)
                break
        for grade, val in grade_dict.items():
            if val == median(class_grade_sum):
                median_grade = grade
        summary.item(1, NUM_QUESTIONS + 4).setText(median_grade)

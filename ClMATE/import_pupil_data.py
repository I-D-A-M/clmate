'''Import pupil data into the database'''

import sqlite3
from openpyxl import load_workbook


def import_pupil_data(sheet_list):
    DB = sqlite3.connect('ClMATE_DB.db')
    addPupil = "insert into cohort (UPN, Name, teaching_set, SEN, PP, KS2Band, KS2lvl, FFT, GCSE_result, ASAlps, ASResult, A2Alps) values (?,?,?,?,?,?,?,?,?,?,?,?)"

    for s in sheet_list:
        # -- Load each workbook as an iterator based openpyxl object
        #    and select the first sheet of each for copying.
        workbook = load_workbook(s, use_iterators=True)
        sheet = workbook.active
        testCohortList = []
        first_row = True
        for row in sheet.iter_rows():
            # -- Skip the header fields in the worksheet
            if first_row:
                first_row = False
            else:
                currentPupil = []
                for cell in row:
                    currentPupil.append(cell.value)
                testCohortList.append(tuple(currentPupil))
        testCohort = tuple(testCohortList)
        # -- Import pupil data into the database
        for pupil in testCohort:
            DB.execute(addPupil, pupil)
        DB.commit()

        a = input('any key to close')

if __name__ == '__main__':
    import_pupil_data(['y9maths.xlsx', 'y10maths.xlsx'])

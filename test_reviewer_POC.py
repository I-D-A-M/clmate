import sqlite3
from openpyxl import Workbook
from collections import namedtuple, OrderedDict


class Reviewer():
    '''
    This Class captures the the data and operations required for performing post
    assessment analysis.
    '''
    def __init__(self, teaching_set, DBname):
        '''
        Grab ALL stored results for the chosen class and improve the storage for later
        use and analysis.
        '''
        self.teaching_set = teaching_set
        self.DBname = DBname

        with sqlite3.connect(DBname) as DB:
            query = "SELECT * FROM results WHERE teaching_set = ?"
            # -- Format is: aID; UPN; teaching_set; aName; qNum; pMark
            res = DB.execute(query, (teaching_set,)).fetchall()

            # -- prep question mark data for combining into the result tuples
            self.assessmentMarks = dict()
            self.assessmentIDs = set([r[0] for r in res])
            for a in self.assessmentIDs:
                query = "SELECT qNum, qMark from assessments where aID = ?"
                qScores = DB.execute(query, (a,)).fetchall()
                for q in qScores:
                    key = str(a) + '/' + str(q[0])
                    self.assessmentMarks[key] = q[1]

        # -- Improve readability for main dataset: class_results should NOT be mutated
        result = namedtuple('Result', ['UPN', 'tSet', 'aID', 'qNum', 'mark', 'total'])
        self.class_results = [result(r[1], r[2], r[0], r[4], r[5], self.assessmentMarks[str(r[0]) + '/' + str(r[4])]) for r in res]

        # -- active_results is set every time we normalise the data against a target.
        self.active_results = list()

    def normalise(self, against='class'):
        '''
        Normalise the pupil results versus:
            Their class         -- "class"
            The whole cohort    -- "cohort"
            Same estimate       -- "estimate"
            Same entry point    -- "entryPoint"
        '''
        # -- Reset the active_results list
        self.active_results = list()

        # -- each aID acts as a key to retrieve a list of question means
        detailDict = dict()

        # -- grab mean question scores for each question sat by the class
        for a in self.assessmentIDs:
            aDetails = dict()
            with sqlite3.connect(self.DBname) as DB:
                query = "select count(*) from assessments where aID = ?"
                numQs = DB.execute(query, (a,)).fetchone()[0]

            for qNum in range(1, numQs + 1):
                classMarks = [r.mark for r in self.class_results if r.aID == a and r.qNum == qNum]
                mean = sum(classMarks) / len(classMarks)
                aDetails[qNum] = mean
            detailDict[a] = aDetails

        # -- NOTE: This is where we normalise the pupil mark with respect to the mean
        result = namedtuple('NormalisedResult', ['UPN', 'tSet', 'aID', 'qNum', 'norm'])
        for r in self.class_results:
            norm = (r.mark - detailDict[r.aID][r.qNum]) / r.total
            normRes = result(r.UPN, r.tSet, r.aID, r.qNum, float("{0:.5f}".format(norm)))
            self.active_results.append(normRes)

    def select(self, mode='worst', numPerPupil=3, overall_performance=False):
        '''
        For each pupil and each assessment, identify their best/worst questions.
        Return is a dict of {aID:
                                {UPN: worst results,
                                ...},
                            ...}
        '''
        assert((mode == 'worst') or (mode == 'best'))
        # -- fetch pupil UPNs
        with sqlite3.connect(self.DBname) as DB:
            query = "SELECT UPN from cohort where teaching_set = ?"
            UPNs = DB.execute(query, (self.teaching_set,)).fetchall()

        classPerformance = dict()
        titleDict = dict()
        for aID in self.assessmentIDs:
            titleDict[aID] = dict()
            with sqlite3.connect(self.DBname) as DB:
                query = "SELECT qNum, qTitle FROM assessments where aID = ?"
                qTitles = DB.execute(query, (aID,)).fetchall()
                for q in qTitles:
                    titleDict[aID][q[0]] = q[1]

            pupilPerformance = dict()
            for UPN in UPNs:
                performance = [r for r in self.active_results if r.aID == aID and r.UPN == UPN[0]]
                performance = sorted(performance, key=lambda x: x.norm)
                if mode == 'worst':
                    perf = performance[:numPerPupil]
                else:
                    perf = performance[-numPerPupil:]
                readable_results = [(titleDict[r.aID][r.qNum], r.norm) for r in perf]

                with sqlite3.connect(self.DBname) as DB:
                    query = "SELECT Name FROM cohort WHERE UPN = ?"
                    name = DB.execute(query, (UPN[0],)).fetchone()[0]

                pupilPerformance[name] = readable_results
                # -- OrderedDict to maintain register order of pupils
                classPerformance[aID] = OrderedDict(sorted(pupilPerformance.items()))

        self.output = classPerformance
        if overall_performance:
            self.output = self.calc_overall_performance(mode, numPerPupil)

    def calc_overall_performance(self, mode, numPerPupil):
        ''' Find the best/worst questions for all assessments'''
        all_results = dict()
        for aID, results in self.output.items():
            for pupil, performance in results.items():
                if pupil not in all_results:
                    all_results[pupil] = performance
                else:
                    for p in performance:
                        all_results[pupil].append(p)

        sortedRes = OrderedDict(sorted(all_results.items()))
        for pupil in sortedRes:
            sortedRes[pupil] = sorted(sortedRes[pupil], key=lambda x: x[1])
            QWC_found = False
            for i, q in enumerate(sortedRes[pupil]):
                if 'QWC' in q[0] and not QWC_found:
                    # Rename QWC questions as targets
                    sortedRes[pupil][i] = ('QWC: Identify the problem topic and select an appropriate strategy', q[1])
                    QWC_found = True
                elif 'QWC' in q[0] and QWC_found:
                    # Remove the duplicate
                    sortedRes[pupil].pop(i)

            if mode == 'worst':
                sortedRes[pupil] = sortedRes[pupil][:numPerPupil]
            else:
                sortedRes[pupil] = sortedRes[pupil][-numPerPupil:]

        return sortedRes

    def write_to_txt(self, filename="output.txt", allAssessments=True):
        with open(filename, "w") as f:
            f.write("Name,Question 1,Question 2,Question 3")
            if not allAssessments:
                for assessment in self.output:
                    with sqlite3.connect(self.DBname) as DB:
                        query = "SELECT aName FROM assessments WHERE aID = ?"
                        aName = DB.execute(query, (assessment,)).fetchone()[0]
                    f.write("Assessment: {}".format(aName))
                    for pupil in self.output[assessment]:
                        name = pupil.split()
                        f.write('\n' + name[1] + ' ' + name[0] + '\n')
                        for Q in self.output[assessment][pupil]:
                            # .xlsx is a CSV format so get rid of any commmas!
                            cleanedTopic = Q[0].replace(',', '').replace(';', '')
                            # Also remove any tabs or newlines to protect formatting
                            f.write("{},".format(cleanedTopic.strip()))
                    f.write('\n')
            else:
                for pupil in self.output:
                    name = pupil.split()
                    f.write('\n' + name[1] + ' ' + name[0])
                    for Q in self.output[pupil]:
                        # .xlsx is a CSV format so get rid of any commmas!
                        cleanedTopic = Q[0].replace(',', '').replace(';', '')
                        # Also remove any tabs or newlines to protect formatting
                        f.write("{},".format(cleanedTopic.strip()))

    def write_to_xlsx(self, filename="output.xlsx", allAssessments=True):
        wb = Workbook()
        # This will create a new sheet --> ws = wb.create_sheet()
        ws = wb.active
        # Set headings
        headings = ["Name", "Question 1", "Question 2", "Question 3"]
        for col in range(4):
            heading = ws.cell(row=0, column=(col))
            heading.value = headings[col]

        if not allAssessments:
            print("Idividual assessments not yet implemented! Try again...")
        else:
            for row, pupil in enumerate(self.output):
                name = pupil.split()
                nameCell = ws.cell(row=row + 1, column=(0))
                nameCell.value = name[1] + ' ' + name[0].replace(',', '')

                for col, Q in enumerate(self.output[pupil]):
                    # .xlsx is a CSV format so get rid of any commmas and semicolons
                    cleanedTopic = Q[0].replace(',', '').replace(';', '')
                    # Also remove any tabs or newlines to protect formatting
                    questionCell = ws.cell(row=row + 1, column=(col + 1))
                    questionCell.value = "{}".format(cleanedTopic.strip())
        wb.save(filename)

if __name__ == '__main__':
    '''
    Generate the target stickers for each class and find the lowest performing questions
    per tier.
    '''
    y11 = ['11-PG', '11-DBY', '11-HH', '11-CCL', '11-AR', '11-AK', '11-IM', '11-MD']

    for tSet in y11:
        r = Reviewer(teaching_set=tSet, DBname="ClMATE_DB.db")
        r.normalise(against='class')
        r.select(mode='worst', numPerPupil=3, overall_performance=True)
        r.write_to_xlsx(tSet + ".xlsx")
